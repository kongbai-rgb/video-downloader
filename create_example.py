import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "视频链接"

# 设置标题
ws['A1'] = "视频链接列表"
ws['A1'].font = Font(size=16, bold=True)
ws['A1'].alignment = Alignment(horizontal='center')
ws.merge_cells('A1:B1')

# 设置列标题
ws['A3'] = "序号"
ws['B3'] = "视频链接"
ws['C3'] = "备注"

# 设置列标题样式
header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
for cell in ['A3', 'B3', 'C3']:
    ws[cell].font = Font(bold=True)
    ws[cell].fill = header_fill
    ws[cell].alignment = Alignment(horizontal='center')

# 添加示例数据
example_links = [
    ("1", "https://www.youtube.com/watch?v=dQw4w9WgXcQ", "示例YouTube视频"),
    ("2", "https://www.instagram.com/p/CZB2AXXXXXZ/", "示例Instagram视频"),
    ("3", "https://youtu.be/XXXXXXXXXXX", "YouTube短链接示例"),
    ("4", "", "在这里添加你的链接"),
    ("5", "", ""),
]

# 填充数据
for idx, (num, link, note) in enumerate(example_links, start=4):
    ws[f'A{idx}'] = num
    ws[f'B{idx}'] = link
    ws[f'C{idx}'] = note
    ws[f'A{idx}'].alignment = Alignment(horizontal='center')

# 调整列宽
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 30

# 添加使用说明
ws['A10'] = "使用说明："
ws['A10'].font = Font(bold=True)
ws['A11'] = "1. 在B列填入要下载的视频链接"
ws['A12'] = "2. 支持YouTube和Instagram链接"
ws['A13'] = "3. 每行一个链接"
ws['A14'] = "4. 保存文件后在下载器中导入"

# 保存文件
wb.save("批量下载示例.xlsx")
print("示例Excel文件已创建：批量下载示例.xlsx")